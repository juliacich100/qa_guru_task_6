import zipfile
from zipfile import ZipFile
import os
import shutil
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

my_archive = zipfile.ZipFile('my_archive.zip', 'w')
my_archive.write('file.csv')
my_archive.write('sky.pdf')
my_archive.write('checklist.xlsx')
my_archive.close()

os.mkdir('resources')

shutil.move('my_archive.zip', 'resources')
with ZipFile('resources/my_archive.zip', 'r') as f:
    f.extractall()
os.remove('resources/my_archive.zip')

def test_scv_file():
    with open('file.csv') as csvfile:
        csvfile = csv.reader(csvfile)
        rowcounter = 0
        for r in csvfile:
            rowcounter += 1
#       print(counter)
        assert rowcounter == 5


def test_pdf_file():
    reader = PdfReader('sky.pdf')
    page = reader.pages[0]
    pdf_text = page.extractText()
    assert 'Warm up' in pdf_text


def test_xlsx_file():
    workbook = load_workbook('checklist.xlsx')
    sheet = workbook.active
    assert sheet.cell(row=6, column=2).value == 'Закрытие формы по клику на Close или вне формы'


shutil.rmtree('resources')